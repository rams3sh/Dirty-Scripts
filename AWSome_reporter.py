import json
from flatten_json import flatten			
import re
import sqlite3
import sys
import time
import os
from openpyxl.workbook import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl import load_workbook
from tqdm import tqdm
#Read the json data from a file into a variable
with open(sys.argv[1]) as f:
	data = json.load(f)

#Flattening the JSON data
normalized_data=flatten(data)
tempsorted_columns=normalized_data.keys()
tempsorted_columns.sort()
columns=[]

#Finding Unique Columns
temp=[]
for i in tempsorted_columns:
	temp=re.search(r'_\d+$', i)
	if temp is not None:
		ref=i.rstrip('1234567890_')
	else:
		ref=i
	ref=re.sub("_\d+_","$#$",ref)
	ref=ref.replace("_","_#_").replace("$","_")
	if ref not in columns:
		columns.append(ref)
	else:
		pass
columns.sort()
temp=columns
#Adding # as one of the columns which determines the record number under each sub category (This is not AWS given value
# But internal reference value )
temp2=[]
for i in columns:
	temp2=[]
	hash_finder=re.compile("_#")
	for m in hash_finder.finditer(i):
		temp2.append(m.start())
	for m in temp2:
		if i.__getslice__(0,m+2) in columns:
			pass
		else:
			temp.append(i.__getslice__(0,m+2))
temp.sort()
columns=temp

#Creating a Database and a table
conn = sqlite3.connect('temp.sqlite')
c=conn.cursor()
table_name="v"
temp=[]
for i in range(len(columns)):
	if columns[i].endswith("#"):
		temp.append('INT')
	else:
		temp.append('TEXT')
query="CREATE TABLE IF NOT EXISTS `"+table_name+"` ("
for i in range(len(columns)):
	query+="`"+columns[i]+"` "+temp[i]+" ,"
query=query.strip(",")+");"
c.execute(query)

reverse_sorted_keys_by_length=normalized_data.keys()
reverse_sorted_keys_by_length.sort(key=lambda item: (-len(item), item))

count=0
for i in tqdm(reverse_sorted_keys_by_length, total=len(reverse_sorted_keys_by_length)):
	#print count
	count+=1
	temp=re.search(r'_\d+$', i)
	if temp is not None:
		ref=i.rstrip('1234567890_')
	else:
		ref=i
	temp=ref
	ref=re.sub("_\d+_","$#$",ref)
	ref=ref.replace("_","_#_").replace("$","_")
	temp=temp.replace("_","$")
	temp=re.sub("(?<=[a-zA-Z])\$(?![0-9])","$0$",temp)
	hash_finder=re.compile("_#")
	temp2=[]
	for m in hash_finder.finditer(ref):
		temp2.append(m.start())
	
	applicable_control_and_values=[]
	
	control_finder=re.compile("_#_")
	value_finder=re.compile("\d+(?<!\$)")
	for m,n in zip(control_finder.finditer(ref),value_finder.finditer(temp)):
		applicable_control_and_values.append(ref.__getslice__(0,m.start()+2))
		applicable_control_and_values.append(n.group())
	applicable_control_and_values.append(ref)
	applicable_control_and_values.append(normalized_data[i])
	
	query=""
	
	query="select `"+applicable_control_and_values[len(applicable_control_and_values)-2]+"` from `"+table_name+"` where `"
	for m in range(0,len(applicable_control_and_values)-2,2):
		query+=applicable_control_and_values[m]+"` = \'"+str(applicable_control_and_values[m+1])+"\'"
		if m < len(applicable_control_and_values)-4:
				query+=" and `"
	query+=";"
	response=c.execute(query).fetchall()
	record=dict.fromkeys(columns, "N/A")
	
	if len(response)==0:
		for m in range(0,len(applicable_control_and_values),2):
			record[applicable_control_and_values[m]]=str(applicable_control_and_values[m+1])
		col='`'+'`, `'.join(record.keys())+'`'
		placeholders = '\''+'\',\''.join(record.values())+"\'"
		query='INSERT INTO '+table_name+' (%s) VALUES (%s)' % (col, placeholders)
		c.execute(query)
	else:
		if (u'N/A',) in response:
			
			if response.count((u'N/A',)) > 1:
				m = re.search(r'\d+$', i)
				if m is not None:
					query="update "+table_name+" set `"+applicable_control_and_values[len(applicable_control_and_values)-2]+"` = \'"+str(applicable_control_and_values[len(applicable_control_and_values)-1])+"\' where `"
					for m in range(0,len(applicable_control_and_values)-2,2):
						query+=applicable_control_and_values[m]+"` = \'"+str(applicable_control_and_values[m+1])+"\'"
						if m < len(applicable_control_and_values)-4:
							query+=" and `"
					query+=" and `"+applicable_control_and_values[len(applicable_control_and_values)-2]+"` = 'N/A' and rowid in (select min(rowid) from `"+table_name+"` where `"
					for m in range(0,len(applicable_control_and_values)-2,2):
						query+=applicable_control_and_values[m]+"` = \'"+str(applicable_control_and_values[m+1])+"\'"
						if m < len(applicable_control_and_values)-4:
							query+=" and `"
					query+=" and `"+applicable_control_and_values[len(applicable_control_and_values)-2]+"` = 'N/A' );"
					c.execute(query)
				else:
					query="update "+table_name+" set `"+applicable_control_and_values[len(applicable_control_and_values)-2]+"` = \'"+str(applicable_control_and_values[len(applicable_control_and_values)-1])+"\' where `"
					for m in range(0,len(applicable_control_and_values)-2,2):
						query+=applicable_control_and_values[m]+"` = \'"+str(applicable_control_and_values[m+1])+"\'"
						if m < len(applicable_control_and_values)-4:
							query+=" and `"
					query+="and `"+applicable_control_and_values[len(applicable_control_and_values)-2]+"` = 'N/A';"
					c.execute(query)
			else:
				query="update "+table_name+" set `"+applicable_control_and_values[len(applicable_control_and_values)-2]+"` = \'"+str(applicable_control_and_values[len(applicable_control_and_values)-1])+"\' where `"
				for m in range(0,len(applicable_control_and_values)-2,2):
					query+=applicable_control_and_values[m]+"` = \'"+str(applicable_control_and_values[m+1])+"\'"
					if m < len(applicable_control_and_values)-4:
						query+=" and `"
				query+="and `"+applicable_control_and_values[len(applicable_control_and_values)-2]+"` = 'N/A';"
				c.execute(query)
				
		else:
			for m in range(0,len(applicable_control_and_values),2):
				record[applicable_control_and_values[m]]=str(applicable_control_and_values[m+1])
			col='`'+'`, `'.join(record.keys())+'`'
			placeholders = '\''+'\',\''.join(record.values())+"\'"
			query='INSERT INTO '+table_name+' (%s) VALUES (%s)' % (col, placeholders)
			c.execute(query)
conn.commit()

#Reporting in Excel 

query="select * from `"+table_name+"` order by "
for i in columns:
	if i.endswith("#"):
		query+="`"+i+"`,"
query=query.strip(",")+";"
queue=c.execute(query)

max=0
for i in columns:
	if max > i.count("_"):
		pass
	else:
		max=i.count("_")

wb = Workbook()
dest_filename = sys.argv[2]+'.xlsx'
ws1 = wb.active
eycolourFill = PatternFill(start_color='FFC000',end_color='FFC000',fill_type='solid')
font = Font(name='Calibri',size=11,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000')
thin_border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))

max=0
temp=[]
for i in range(len(columns)):
	if columns[i].__contains__("_"):
		k=columns[i].split("_")
		if max < len(k):
			max=len(k)
		for l,t in zip(k,range(len(k))):
			ws1.cell(row=t+1,column=i+1).value=l
			ws1.cell(row=t+1,column=i+1).fill = eycolourFill
			ws1.cell(row=t+1,column=i+1).font=font
			ws1.cell(row=t+1,column=i+1).border=thin_border
			
	else:	
		ws1.cell(row=1,colum=i+1).value=columns[i]
		ws1.cell(row=1,colum=i+1).fill=eycolourFill
		ws1.cell(row=1,colum=i+1).font=font
		ws1.cell(row=1,colum=i+1).border=thin_border
wb.save(filename = dest_filename)

wbr=load_workbook(sys.argv[2]+".xlsx",read_only=True)
ws2 = wb.active


for i in range(1,max,2):
	temp1=0
	temp=ws2.cell(i,1).value
	start=1
	for j in range(1,len(columns)):
		if temp == ws2.cell(i,j).value :
			if temp == None:
				start=j
		else:
			temp1=1
			ws1.merge_cells(start_row=i, start_column=start, end_row=i, end_column=j-1)
			if ws2.cell(i+1,start).value== "#":
				ws1.merge_cells(start_row=i+1, start_column=start, end_row=i+1, end_column=j-1)
			temp=ws2.cell(i,j).value
			wb.save(filename = dest_filename)
			start=j
	if temp1 == 0:
		ws1.merge_cells(start_row=i, start_column=start, end_row=i, end_column=j+1)
		if ws2.cell(i+1,start).value== "#":
			ws1.merge_cells(start_row=i+1, start_column=start, end_row=i+1, end_column=j+1)
				
		
x=max
y=0	
while True:
    rows = queue.fetchmany(100)
    if not rows: 	
		break
    for row in rows:
		y=0
		for j in row:
			ws1.cell(row=x+1,column=y+1).value=j
			ws1.cell(row=x+1,column=y+1).border=thin_border
			y+=1
		x+=1
wb.save(filename = dest_filename)	
conn.close()






